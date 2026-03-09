"""
CSV Cleaner Telegram Bot — with conversational pivot editing
Send messy CSV data or a .csv file, get back a Numbers-style pivot table.
Then chat with the bot to ask questions or edit the table.
"""

import os
import io
import json
import asyncio
import logging
import pandas as pd
from telegram import Bot, Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TOKEN   = os.environ.get("TELEGRAM_BOT_TOKEN", "")
CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")

anthropic_client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

# Per-user session: { user_id: { "csv": str, "pivot_config": dict, "history": [...] } }
sessions: dict = {}

# Apple Numbers colors
NUMBERS_HEADER_BG  = "1B6EC2"
NUMBERS_HEADER_FG  = "FFFFFF"
NUMBERS_ROW_ALT    = "EEF4FB"
NUMBERS_ROW_WHITE  = "FFFFFF"
NUMBERS_TOTAL_BG   = "D0E4F5"
NUMBERS_TOTAL_FG   = "0D3C6B"
NUMBERS_BORDER     = "C5D8EE"
NUMBERS_TITLE_BG   = "0D3C6B"
NUMBERS_TITLE_FG   = "FFFFFF"


# ── send_file helper (for use from data_cleaner.py) ──────────────────────────

def send_file(file):
    bot = Bot(token=TOKEN)
    async def _send():
        with open(file, "rb") as f:
            await bot.send_document(chat_id=CHAT_ID, document=f)
    asyncio.run(_send())


# ── Claude calls ──────────────────────────────────────────────────────────────

def clean_and_plan_pivot(messy_data: str) -> dict:
    """Clean messy data and suggest a pivot config. Returns {csv, pivot}."""
    msg = anthropic_client.messages.create(
        model="claude-3-haiku-20240307",
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": f"""You are a data analyst. Do two things:

1. Clean this messy data into proper CSV:
   - Title Case names, fix delimiters (|/-/tabs → commas)
   - Plain numbers (no $, no k: 50k → 50000)
   - Add header row if missing, remove empty rows, trim whitespace

2. Suggest the best pivot table config (like Apple Numbers):
   - index: categorical row-grouping columns
   - columns: one column-grouping field, or null
   - values: numeric columns to aggregate
   - aggfunc: "sum", "mean", or "count"

Return ONLY valid JSON, nothing else:
{{
  "csv": "col1,col2,...\\nval1,val2,...",
  "pivot": {{
    "index": ["ColA"],
    "columns": "ColB",
    "values": ["ColC"],
    "aggfunc": "sum"
  }}
}}

Data:
{messy_data}"""
        }]
    )
    raw = msg.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def chat_about_table(csv_data: str, pivot_config: dict, history: list, user_message: str) -> dict:
    """
    Full data editing + pivot conversation.
    Returns {
      "answer": str,
      "new_csv": str|null,       -- updated CSV if data changed
      "new_pivot": dict|null,    -- updated pivot config if changed
      "filter": str|null,        -- pandas query string to filter rows
      "sort_by": str|null,       -- column name to sort by
      "sort_asc": bool           -- sort direction
    }
    """
    df = pd.read_csv(io.StringIO(csv_data.strip()), on_bad_lines='skip')
    full_data = df.to_csv(index=False)
    columns = list(df.columns)

    system = f"""You are a data analyst helping the user edit their data and pivot table.

Current columns: {columns}

Full data (CSV):
{full_data}

Current pivot:
- Rows: {pivot_config.get('index')}
- Columns: {pivot_config.get('columns')}
- Values: {pivot_config.get('values')} ({pivot_config.get('aggfunc')})

You can make ANY of these changes:
1. Add a new column (even with user-provided values or calculated from existing ones)
2. Add new rows
3. Edit existing values
4. Delete rows or columns
5. Change pivot grouping, values, or aggregation
6. Filter rows (e.g. only show certain dates)
7. Sort by a column
8. Add calculated columns (e.g. Tax = Pay * 0.2)

Always return valid JSON only:
{{
  "answer": "What you did, explained simply",
  "new_csv": null,
  "new_pivot": null,
  "filter": null,
  "sort_by": null,
  "sort_asc": true
}}

Rules for each field:
- new_csv: full updated CSV string if data was added/edited/deleted. Include ALL rows and headers.
- new_pivot: ONLY {{ "index": [...], "columns": "..." or null, "values": [...], "aggfunc": "sum"/"mean"/"count" }} — never put table data here
- filter: a pandas DataFrame.query() string e.g. "Pay > 100" or "Week == 'March 8'"
- sort_by: column name string to sort by, or null
- sort_asc: true for ascending, false for descending
- Set fields to null if not changing them"""

    messages = history + [{"role": "user", "content": user_message}]

    msg = anthropic_client.messages.create(
        model="claude-3-5-haiku-20241022",
        max_tokens=4096,
        system=system,
        messages=messages
    )

    raw = msg.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {"answer": raw, "new_csv": None, "new_pivot": None, "filter": None, "sort_by": None, "sort_asc": True}


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_pivot_excel(csv_data: str, pivot_config: dict) -> io.BytesIO:
    df = pd.read_csv(io.StringIO(csv_data.strip()), on_bad_lines='skip')
    df.columns = [str(c).strip() for c in df.columns]

    valid_values = []
    for col in pivot_config.get("values", []):
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(r'[$,]', '', regex=True).str.strip(),
                errors='coerce'
            )
            valid_values.append(col)
    if not valid_values:
        valid_values = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    pivot_config["values"] = valid_values

    pivot_config["index"] = [c for c in pivot_config.get("index", []) if c in df.columns]
    if not pivot_config["index"]:
        pivot_config["index"] = [df.columns[0]]

    if pivot_config.get("columns") and pivot_config["columns"] not in df.columns:
        pivot_config["columns"] = None

    pivot_kwargs = {
        "index": pivot_config["index"],
        "values": pivot_config["values"],
        "aggfunc": pivot_config["aggfunc"],
        "margins": True,
        "margins_name": "Total",
    }
    if pivot_config.get("columns"):
        pivot_kwargs["columns"] = pivot_config["columns"]

    pivot = pd.pivot_table(df, **pivot_kwargs)
    if isinstance(pivot.columns, pd.MultiIndex):
        pivot.columns = [" · ".join(str(c) for c in col if str(c) != 'nan').strip() for col in pivot.columns]
    pivot = pivot.reset_index()
    raw_rows = [list(df.columns)] + df.values.tolist()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pivot Table"

    def border():
        s = Side(style='thin', color=NUMBERS_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    num_cols = len(pivot.columns)
    num_rows = len(pivot)

    # Title bar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    tc = ws.cell(row=1, column=1)
    agg_label = {"sum": "Sum", "mean": "Average", "count": "Count"}.get(pivot_config["aggfunc"], "Summary")
    tc.value = f"Pivot Table  —  {agg_label} of {', '.join(pivot_config['values'])}"
    tc.font = Font(bold=True, color=NUMBERS_TITLE_FG, size=13, name="Helvetica Neue")
    tc.fill = fill(NUMBERS_TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Header row
    for ci, col_name in enumerate(pivot.columns, 1):
        cell = ws.cell(row=2, column=ci, value=str(col_name))
        cell.font = Font(bold=True, color=NUMBERS_HEADER_FG, size=11, name="Helvetica Neue")
        cell.fill = fill(NUMBERS_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border()
    ws.row_dimensions[2].height = 22

    # Data rows
    for ri, (_, row) in enumerate(pivot.iterrows(), 3):
        is_total = str(row.iloc[0]) == "Total"
        is_alt   = (ri % 2 == 0)
        if is_total:
            rf, fc, bold = fill(NUMBERS_TOTAL_BG), NUMBERS_TOTAL_FG, True
        elif is_alt:
            rf, fc, bold = fill(NUMBERS_ROW_ALT), "000000", False
        else:
            rf, fc, bold = fill(NUMBERS_ROW_WHITE), "000000", False

        for ci, value in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            if pd.isna(value):
                cell.value = "—"
            elif isinstance(value, float):
                cell.value = round(value, 2) if pivot_config["aggfunc"] == "mean" else (int(value) if value == int(value) else round(value, 2))
                cell.number_format = '#,##0.00' if pivot_config["aggfunc"] == "mean" else '#,##0'
            else:
                # Clean up ugly float-formatted strings like "March 8 2026.0"
                s = str(value)
                cell.value = s[:-2] if s.endswith(".0") else s
            cell.font = Font(bold=bold, color=fc, size=11, name="Helvetica Neue")
            cell.fill = rf
            cell.border = border()
            cell.alignment = Alignment(
                horizontal="right" if ci > len(pivot_config["index"]) else "left",
                vertical="center",
                indent=0 if ci > len(pivot_config["index"]) else 1
            )
        ws.row_dimensions[ri].height = 20

    # Column widths
    for ci in range(1, num_cols + 1):
        col_letter = get_column_letter(ci)
        max_len = max((len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(1, num_rows + 4)), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A3"

    # Raw data sheet
    ws2 = wb.create_sheet(title="Cleaned Data")
    for ri, row in enumerate(raw_rows, 1):
        for ci, value in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=str(value).strip() if ri > 1 else value)
            if ri == 1:
                cell.font = Font(bold=True, color=NUMBERS_HEADER_FG)
                cell.fill = fill(NUMBERS_HEADER_BG)
                cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Helpers ───────────────────────────────────────────────────────────────────

def is_new_data(text: str) -> bool:
    """Heuristic: if text has multiple lines with delimiters, treat as new CSV data."""
    lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
    if len(lines) < 2:
        return False
    delimiters = sum(1 for l in lines if any(c in l for c in [',', '|', '\t', ' - ']))
    return delimiters >= len(lines) * 0.5


def is_file_request(text: str) -> bool:
    """Detect when the user just wants the file resent."""
    keywords = ["give me the", "send the", "send me the", "the file", "the table",
                "resend", "re-send", "get the file", "download"]
    t = text.lower().strip()
    return any(k in t for k in keywords)


def is_valid_pivot_config(p) -> bool:
    """Check that new_pivot has the right structure, not table data."""
    if not isinstance(p, dict):
        return False
    if "index" not in p or "values" not in p or "aggfunc" not in p:
        return False
    if not isinstance(p.get("index"), list) or not isinstance(p.get("values"), list):
        return False
    if p.get("aggfunc") not in ("sum", "mean", "count"):
        return False
    return True


def pivot_caption(pivot_config: dict) -> str:
    agg = {"sum": "Sum", "mean": "Average", "count": "Count"}.get(pivot_config["aggfunc"], pivot_config["aggfunc"])
    return (
        f"✅ *Pivot Table ready!*\n\n"
        f"📊 *Rows:* {', '.join(pivot_config['index'])}\n"
        + (f"📋 *Columns:* {pivot_config['columns']}\n" if pivot_config.get('columns') else "")
        + f"🔢 *Values:* {', '.join(pivot_config['values'])} ({agg})\n\n"
        f"💬 _Ask me anything about the table, or tell me how to change it!_"
    )


# ── Handlers ──────────────────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *Welcome to messydata2!*\n\n"
        "Send me messy data or a `.csv` file and I'll build a *Numbers-style pivot table*.\n\n"
        "After that, just chat with me — ask questions about the data or tell me how to edit the table! ✨",
        parse_mode='Markdown'
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "*📖 Help*\n\n"
        "1️⃣ Send messy data or a `.csv` file → get a pivot table\n"
        "2️⃣ Then ask questions or request changes:\n"
        "   • _\"What's the highest salary?\"_\n"
        "   • _\"Group by Department instead\"_\n"
        "   • _\"Show averages not sums\"_\n"
        "   • _\"Add Region as columns\"_\n\n"
        "*/start* — Welcome  |  */reset* — Start over",
        parse_mode='Markdown'
    )


async def reset_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sessions.pop(user_id, None)
    await update.message.reply_text("🔄 Session cleared. Send me new data to start fresh!")


async def process_new_data(update: Update, messy_data: str, output_filename: str):
    user_id = update.effective_user.id
    processing_msg = await update.message.reply_text("⏳ Cleaning data and building pivot table...")

    try:
        result = await asyncio.get_event_loop().run_in_executor(
            None, clean_and_plan_pivot, messy_data
        )
        clean_csv    = result["csv"]
        pivot_config = result["pivot"]
        excel_file   = build_pivot_excel(clean_csv, pivot_config)

        excel_bytes = excel_file.read()

        # Save session
        sessions[user_id] = {
            "csv": clean_csv,
            "pivot_config": pivot_config,
            "history": [],
            "last_excel": excel_bytes,
            "last_filename": output_filename,
        }

        await processing_msg.delete()
        await update.message.reply_document(
            document=io.BytesIO(excel_bytes),
            filename=output_filename,
            caption=pivot_caption(pivot_config),
            parse_mode='Markdown'
        )

    except Exception as e:
        logger.error(f"Error processing data: {e}")
        await processing_msg.edit_text(f"❌ Error: {str(e)}\n\nPlease try again.")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text    = update.message.text

    # If it looks like new CSV data, process it fresh
    if is_new_data(text):
        await process_new_data(update, text, "pivot_table.xlsx")
        return

    # If there's an active session, treat as conversation
    if user_id in sessions:
        session = sessions[user_id]

        # Resend the file if user just wants the file/table
        if is_file_request(text) and session.get("last_excel"):
            await update.message.reply_document(
                document=io.BytesIO(session["last_excel"]),
                filename=session.get("last_filename", "pivot_table.xlsx"),
                caption=pivot_caption(session["pivot_config"]),
                parse_mode='Markdown'
            )
            return

        thinking_msg = await update.message.reply_text("💭 Thinking...")

        try:
            response = await asyncio.get_event_loop().run_in_executor(
                None,
                lambda: chat_about_table(
                    session["csv"],
                    session["pivot_config"],
                    session["history"],
                    text
                )
            )

            answer    = response.get("answer", "")
            new_csv   = response.get("new_csv")
            new_pivot = response.get("new_pivot")
            filter_q  = response.get("filter")
            sort_by   = response.get("sort_by")
            sort_asc  = response.get("sort_asc", True)

            # Update conversation history
            session["history"].append({"role": "user",      "content": text})
            session["history"].append({"role": "assistant", "content": answer})

            await thinking_msg.delete()
            await update.message.reply_text(answer)

            # Apply data changes
            rebuild = False

            if new_csv:
                session["csv"] = new_csv
                rebuild = True

            # Apply filter
            if filter_q:
                try:
                    df = pd.read_csv(io.StringIO(session["csv"].strip()), on_bad_lines='skip')
                    df = df.query(filter_q)
                    session["csv"] = df.to_csv(index=False)
                    rebuild = True
                except Exception as fe:
                    await update.message.reply_text(f"⚠️ Couldn't apply filter: {fe}")

            # Apply sort
            if sort_by:
                try:
                    df = pd.read_csv(io.StringIO(session["csv"].strip()), on_bad_lines='skip')
                    if sort_by in df.columns:
                        df = df.sort_values(sort_by, ascending=sort_asc)
                        session["csv"] = df.to_csv(index=False)
                        rebuild = True
                except Exception as se:
                    await update.message.reply_text(f"⚠️ Couldn't sort: {se}")

            # Apply pivot config change
            if new_pivot and is_valid_pivot_config(new_pivot):
                session["pivot_config"] = new_pivot
                rebuild = True
            elif new_pivot and not is_valid_pivot_config(new_pivot):
                df = pd.read_csv(io.StringIO(session["csv"].strip()), on_bad_lines='skip')
                await update.message.reply_text(
                    f"⚠️ Couldn't update pivot config — invalid structure.\n"
                    f"Available columns: *{', '.join(df.columns.tolist())}*",
                    parse_mode='Markdown'
                )

            # Rebuild and send Excel if anything changed
            if rebuild:
                excel_file = build_pivot_excel(session["csv"], session["pivot_config"])
                excel_bytes = excel_file.read()
                session["last_excel"] = excel_bytes
                session["last_filename"] = "pivot_updated.xlsx"
                await update.message.reply_document(
                    document=io.BytesIO(excel_bytes),
                    filename="pivot_updated.xlsx",
                    caption=pivot_caption(session["pivot_config"]),
                    parse_mode='Markdown'
                )

        except Exception as e:
            logger.error(f"Chat error: {e}")
            await thinking_msg.edit_text(f"❌ Error: {str(e)}")

    else:
        # No session yet — treat short text as potential data attempt
        await update.message.reply_text(
            "Send me some data first! Paste rows or upload a `.csv` file and I'll build a pivot table. "
            "Then you can ask me questions about it. 📊"
        )


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    if not document.file_name.lower().endswith('.csv'):
        await update.message.reply_text("Please send a `.csv` file, or paste your data as text!")
        return
    try:
        file       = await context.bot.get_file(document.file_id)
        file_bytes = await file.download_as_bytearray()
        messy_data = file_bytes.decode('utf-8')
        output_name = document.file_name.replace('.csv', '_pivot.xlsx')
        await process_new_data(update, messy_data, output_name)
    except Exception as e:
        logger.error(f"Error processing file: {e}")
        await update.message.reply_text(f"❌ Error: {str(e)}\n\nPlease try again.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not TOKEN:
        print("❌ TELEGRAM_BOT_TOKEN not set!")
        return
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("❌ ANTHROPIC_API_KEY not set!")
        return

    print("🤖 Starting messydata2 bot...")
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help",  help_command))
    app.add_handler(CommandHandler("reset", reset_command))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
